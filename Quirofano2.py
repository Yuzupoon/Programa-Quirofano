import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import Workbook, load_workbook
import os

ARCHIVO_EXCEL = "checklist_pacientes.xlsx"
pacientes = {}

# --- Funciones Excel ---
def guardar_excel():
    if os.path.exists(ARCHIVO_EXCEL):
        wb = load_workbook(ARCHIVO_EXCEL)
        ws = wb.active
        ws.delete_rows(2, ws.max_row)  # Borra datos viejos, deja encabezado
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Checklist"
        ws.append(["Paciente", "Ítem", "Estado"])
    
    for paciente, items in pacientes.items():
        for item in items:
            estado = "✔️" if item["estado"] else "❌"
            ws.append([paciente, item["item"], estado])
    
    wb.save(ARCHIVO_EXCEL)

def cargar_excel():
    if os.path.exists(ARCHIVO_EXCEL):
        wb = load_workbook(ARCHIVO_EXCEL)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            paciente, item, estado = row
            if paciente not in pacientes:
                pacientes[paciente] = []
            pacientes[paciente].append({
                "item": item,
                "estado": True if estado == "✔️" else False
            })

# --- Funciones GUI ---
def agregar_paciente():
    nombre = entry_paciente.get().strip()
    item = entry_item.get().strip()
    if nombre and item:
        if nombre not in pacientes:
            pacientes[nombre] = []
        pacientes[nombre].append({"item": item, "estado": False})
        actualizar_lista()
        entry_item.delete(0, tk.END)
        guardar_excel()
    else:
        messagebox.showwarning("Atención", "Debe ingresar paciente e ítem.")

def marcar_item():
    seleccion = lista_items.selection()
    if seleccion:
        paciente, idx = lista_items.item(seleccion[0], "values")[0], int(lista_items.item(seleccion[0], "values")[1])
        pacientes[paciente][idx]["estado"] = True
        actualizar_lista()
        guardar_excel()

def actualizar_lista():
    for row in lista_items.get_children():
        lista_items.delete(row)
    for paciente, items in pacientes.items():
        for i, item in enumerate(items):
            estado = "✔️" if item["estado"] else "❌"
            lista_items.insert("", tk.END, values=(paciente, i, item["item"], estado))

# --- Interfaz gráfica ---
root = tk.Tk()
root.title("Checklist de Pacientes")

frame = tk.Frame(root)
frame.pack(pady=10)

tk.Label(frame, text="Paciente:").grid(row=0, column=0)
entry_paciente = tk.Entry(frame)
entry_paciente.grid(row=0, column=1)

tk.Label(frame, text="Ítem:").grid(row=1, column=0)
entry_item = tk.Entry(frame)
entry_item.grid(row=1, column=1)

btn_agregar = tk.Button(frame, text="Agregar", command=agregar_paciente)
btn_agregar.grid(row=2, column=0, columnspan=2, pady=5)

cols = ("Paciente", "Índice", "Ítem", "Estado")
lista_items = ttk.Treeview(root, columns=cols, show="headings")
for col in cols:
    lista_items.heading(col, text=col)
lista_items.pack(pady=10)

btn_marcar = tk.Button(root, text="Marcar como realizado", command=marcar_item)
btn_marcar.pack(pady=5)

# --- Al iniciar, cargar datos del Excel ---
cargar_excel()
actualizar_lista()

root.mainloop()