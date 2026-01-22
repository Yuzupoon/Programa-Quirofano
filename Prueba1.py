import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import Workbook, load_workbook
import os

ARCHIVO_EXCEL = "checklist_pacientes.xlsx"
pacientes = {}

# --- Funciones Excel ---
def guardar_excel():
    if os.path.exists(ARCHIVO_EXCEL):
        wb = load_workbook(ARCHIVO_EXCEL)
        ws = wb.active
        ws.delete_rows(2, ws.max_row)  # Borra datos viejos, deja encabezado
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Checklist"
        ws.append(["Paciente", "Ítem", "Estado"])
    
    for paciente, items in pacientes.items():
        for item in items:
            estado = "✔️" if item["estado"] else "❌"
            ws.append([paciente, item["item"], estado])
    
    wb.save(ARCHIVO_EXCEL)

def cargar_excel():
    if os.path.exists(ARCHIVO_EXCEL):
        wb = load_workbook(ARCHIVO_EXCEL)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            paciente, item, estado = row
            if paciente not in pacientes:
                pacientes[paciente] = []
            pacientes[paciente].append({
                "item": item,
                "estado": True if estado == "✔️" else False
            })

# --- Ventanas secundarias ---
def ventana_agregar():
    win = tk.Toplevel(root)
    win.title("Agregar paciente/ítem")

    tk.Label(win, text="Paciente:").grid(row=0, column=0)
    entry_paciente = tk.Entry(win)
    entry_paciente.grid(row=0, column=1)

    tk.Label(win, text="Ítem:").grid(row=1, column=0)
    entry_item = tk.Entry(win)
    entry_item.grid(row=1, column=1)

    def agregar():
        nombre = entry_paciente.get().strip()
        item = entry_item.get().strip()
        if nombre and item:
            if nombre not in pacientes:
                pacientes[nombre] = []
            pacientes[nombre].append({"item": item, "estado": False})
            guardar_excel()
            messagebox.showinfo("Éxito", f"Se agregó '{item}' a {nombre}")
            win.destroy()
        else:
            messagebox.showwarning("Atención", "Debe ingresar paciente e ítem.")

    tk.Button(win, text="Guardar", command=agregar).grid(row=2, column=0, columnspan=2, pady=5)

def ventana_buscar():
    win = tk.Toplevel(root)
    win.title("Buscar paciente")

    tk.Label(win, text="Nombre del paciente:").pack(pady=5)
    entry_buscar = tk.Entry(win)
    entry_buscar.pack(pady=5)

    resultado = tk.Text(win, width=50, height=10)
    resultado.pack(pady=5)

    def buscar():
        nombre = entry_buscar.get().strip()
        resultado.delete("1.0", tk.END)
        if nombre in pacientes:
            for i, item in enumerate(pacientes[nombre]):
                estado = "✔️" if item["estado"] else "❌"
                resultado.insert(tk.END, f"{i+1}. {item['item']} - {estado}\n")
        else:
            resultado.insert(tk.END, "Paciente no encontrado.")

    tk.Button(win, text="Buscar", command=buscar).pack(pady=5)

# --- Ventana principal ---
root = tk.Tk()
root.title("Checklist de Pacientes")

btn_agregar = tk.Button(root, text="Agregar", width=20, command=ventana_agregar)
btn_agregar.pack(pady=10)

btn_buscar = tk.Button(root, text="Buscar", width=20, command=ventana_buscar)
btn_buscar.pack(pady=10)

# Cargar datos al iniciar
cargar_excel()

root.mainloop()